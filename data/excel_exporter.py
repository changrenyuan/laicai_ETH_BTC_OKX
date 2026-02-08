# -*- coding: utf-8 -*-
"""
Excel 导出模块
用于将账户信息、持仓、策略等数据导出为 Excel 文件
"""
import sqlite3
from pathlib import Path
from datetime import datetime
from contextlib import contextmanager
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel 导出器"""

    # 数据库字段名常量（避免硬编码）
    FIELD_ID = "id"
    FIELD_TIMESTAMP = "timestamp"
    FIELD_TOTAL_EQUITY_USD = "total_equity_usd"  # 修复：移除多余的 't'，与 persistence.py 保持一致
    FIELD_DETAILS = "details"
    FIELD_INST_ID = "inst_id"
    FIELD_POS_SIDE = "pos_side"
    FIELD_POS = "pos"
    FIELD_AVG_PX = "avg_px"
    FIELD_UNREALIZED_PNL = "unrealized_pnl"
    FIELD_UNREALIZED_PNL_RATIO = "unrealized_pnl_ratio"
    FIELD_LEVERAGE = "leverage"
    FIELD_MARGIN = "margin"
    FIELD_LAST = "last"
    FIELD_STRATEGY_ID = "strategy_id"
    FIELD_STRATEGY_TYPE = "strategy_type"
    FIELD_STATUS = "status"
    FIELD_START_TIME = "start_time"
    FIELD_TOTAL_MARGIN_USD = "total_margin_usd"
    FIELD_TOTAL_CONTRACTS = "total_contracts"
    FIELD_AVG_ENTRY_PRICE = "avg_entry_price"
    FIELD_TP_PRICE = "tp_price"
    FIELD_SL_PRICE = "sl_price"
    FIELD_LIQ_PRICE = "liq_price"
    FIELD_TP_TRIGGER_PX = "tp_trigger_px"
    FIELD_SL_TRIGGER_PX = "sl_trigger_px"
    FIELD_SZ = "sz"
    FIELD_CREATED_TIME = "created_time"
    FIELD_TRIGGERED_TIME = "triggered_time"
    FIELD_TRIGGER_TYPE = "trigger_type"
    FIELD_EXIT_PRICE = "exit_price"
    FIELD_ORDER_ID = "order_id"
    FIELD_SIDE = "side"
    FIELD_ORD_TYPE = "ord_type"
    FIELD_PX = "px"
    FIELD_FILLED_SZ = "filled_sz"
    FIELD_STATE = "state"
    FIELD_SOURCE = "source"
    FIELD_ACTION_TYPE = "action_type"
    FIELD_ACTION_DETAIL = "action_detail"

    def __init__(self, db_path: str = "data/trading_history.db"):
        """
        初始化 Excel 导出器

        :param db_path: 数据库文件路径
        """
        self.db_path = Path(db_path)
        if not self.db_path.exists():
            raise FileNotFoundError(f"数据库文件不存在: {db_path}")

        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row

    def __enter__(self):
        """上下文管理器入口"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口，确保连接关闭"""
        self.close()

    def export_to_excel(self, output_path: str = None) -> str:
        """
        导出所有数据到 Excel 文件

        :param output_path: 输出文件路径，默认为 data/laicai_trading_YYYYMMDD_HHMMSS.xlsx
        :return: Excel 文件路径
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"data/laicai_trading_{timestamp}.xlsx"

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            # 创建工作簿
            wb = Workbook()
            # 删除默认的 Sheet
            wb.remove(wb.active)

            # 导出各个 sheet（独立处理，每个失败不影响其他）
            export_functions = [
                ("账户余额", self._export_balance_sheet),
                ("当前持仓", self._export_positions_sheet),
                ("活跃策略", self._export_active_strategies_sheet),
                ("止盈止损单", self._export_tp_sl_orders_sheet),
                ("订单记录", self._export_orders_sheet),
                ("行为日志", self._export_action_logs_sheet),
            ]

            failed_sheets = []
            for sheet_name, export_func in export_functions:
                try:
                    export_func(wb)
                    logger.debug(f"✅ 成功导出: {sheet_name}")
                except Exception as e:
                    logger.error(f"❌ 导出 {sheet_name} 失败: {e}")
                    failed_sheets.append(sheet_name)

                    # 创建错误提示 sheet
                    try:
                        ws = wb.create_sheet(f"{sheet_name}(错误)")
                        ws.cell(row=1, column=1, value=f"导出失败: {str(e)}")
                        ws.cell(row=2, column=1, value="请检查数据库表结构或联系开发者")
                    except:
                        pass

            # 保存文件
            wb.save(output_path)

            # 返回结果（即使部分失败也返回文件路径）
            if failed_sheets:
                logger.warning(f"⚠️  部分 sheet 导出失败: {', '.join(failed_sheets)}")

            return str(output_path)

        except Exception as e:
            logger.error(f"❌ Excel 导出整体失败: {e}")
            # 确保在异常时仍然关闭连接
            raise e

    def _format_value(self, value):
        """
        格式化值，仅处理 None，保留原始类型以便 openpyxl 识别

        :param value: 原始值
        :return: 格式化后的值
        """
        # 仅处理 None，保留数字类型以便 Excel 识别
        return "" if value is None else value

    def _get_column_width(self, data):
        """
        计算列宽（考虑中文字符宽度）

        :param data: 数据列表
        :return: 列宽字典
        """
        widths = {}
        for row in data:
            for i, value in enumerate(row):
                val = str(value)
                # 计算视觉宽度：中文算2字节，英文算1字节
                width = sum(2 if ord(c) > 127 else 1 for c in val)
                widths[i] = max(widths.get(i, 0), width)
        return widths

    def _format_header_row(self, ws, headers):
        """
        格式化表头行（独立方法，避免重复）

        :param ws: 工作表对象
        :param headers: 表头列表
        """
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def _format_sheet(self, ws, headers, data):
        """
        格式化工作表

        :param ws: 工作表对象
        :param headers: 表头列表
        :param data: 数据列表
        """
        # 先写入表头（无论是否有数据）
        self._format_header_row(ws, headers)

        # 处理空数据情况
        if not data:
            return

        # 写入数据
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                # 格式化值
                formatted_value = self._format_value(value)

                cell = ws.cell(row=row_num, column=col_num, value=formatted_value)
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # 数值格式化
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.number_format = "#,##0.00"

                # 边框
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

        # 自动调整列宽（处理中文显示）
        column_widths = self._get_column_width([headers] + data)
        for col_num, width in column_widths.items():
            # col_num 从 0 开始（enumerate 的索引），需要 +1 转换为 openpyxl 的列索引（A=1, B=2, ...）
            col_letter = get_column_letter(col_num + 1)
            # 给予 1.2 的缓冲系数，确保内容不拥挤
            adjusted_width = min(max(width * 1.2, 12), 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        # 冻结首行
        ws.freeze_panes = "A2"

    def _export_balance_sheet(self, wb):
        """导出账户余额"""
        ws = wb.create_sheet("账户余额")

        cursor = self.conn.cursor()

        try:
            # 先检查表结构，获取实际字段名
            cursor.execute("PRAGMA table_info(account_balance)")
            columns_info = cursor.fetchall()
            available_columns = [col[1] for col in columns_info]

            # 检查是否存在 total_equity_usd 字段
            equity_field = self.FIELD_TOTAL_EQUITY_USD
            if self.FIELD_TOTAL_EQUITY_USD not in available_columns:
                # 尝试其他可能的字段名
                for field in ["total_equity_usdt", "total_equity", "totalEq", "total_usd"]:
                    if field in available_columns:
                        equity_field = field
                        break

                # 如果还是找不到，记录警告
                if equity_field not in available_columns:
                    logger.warning(f"未找到权益字段，可用字段: {available_columns}")

            # 明确指定字段名，使用实际存在的字段
            query_fields = [self.FIELD_ID, self.FIELD_TIMESTAMP, equity_field, self.FIELD_DETAILS]
            query_fields_str = ", ".join(query_fields)

            cursor.execute(f"""
                SELECT {query_fields_str}
                FROM account_balance
                ORDER BY {self.FIELD_TIMESTAMP} DESC
                LIMIT 50
            """)

            rows = cursor.fetchall()

        except sqlite3.OperationalError as e:
            logger.error(f"查询账户余额失败: {e}")
            # 创建空表并写入错误提示
            headers = ["错误信息"]
            data = [[f"数据库字段不匹配: {e}"], ["可用字段: " + str(available_columns)]]
            self._format_sheet(ws, headers, data)
            return

        headers = ["ID", "时间", "总权益(USD)", "详情"]
        data = []

        for row in rows:
            detail = row[self.FIELD_DETAILS] or ""
            # 截断过长的详情（确保 detail 不为 None）
            if len(detail) > 100:
                detail = detail[:100] + "..."

            data.append([
                row[self.FIELD_ID],
                row[self.FIELD_TIMESTAMP],
                row[equity_field],
                detail
            ])

        self._format_sheet(ws, headers, data)

    def _export_positions_sheet(self, wb):
        """导出当前持仓"""
        ws = wb.create_sheet("当前持仓")

        cursor = self.conn.cursor()
        cursor.execute(f"""
            SELECT {self.FIELD_ID}, {self.FIELD_TIMESTAMP}, {self.FIELD_INST_ID}, {self.FIELD_POS_SIDE},
                   {self.FIELD_POS}, {self.FIELD_AVG_PX}, {self.FIELD_UNREALIZED_PNL},
                   {self.FIELD_UNREALIZED_PNL_RATIO}, {self.FIELD_LEVERAGE}, {self.FIELD_MARGIN}, {self.FIELD_LAST}
            FROM positions
            ORDER BY {self.FIELD_TIMESTAMP} DESC
            LIMIT 50
        """)

        rows = cursor.fetchall()

        headers = ["ID", "时间", "交易对", "方向", "持仓量", "均价", "未实现盈亏", "未实现盈亏率", "杠杆", "保证金", "最新价"]
        data = []

        for row in rows:
            data.append([
                row[self.FIELD_ID],
                row[self.FIELD_TIMESTAMP],
                row[self.FIELD_INST_ID],
                row[self.FIELD_POS_SIDE],
                row[self.FIELD_POS],
                row[self.FIELD_AVG_PX],
                row[self.FIELD_UNREALIZED_PNL],
                row[self.FIELD_UNREALIZED_PNL_RATIO],
                row[self.FIELD_LEVERAGE],
                row[self.FIELD_MARGIN],
                row[self.FIELD_LAST]
            ])

        self._format_sheet(ws, headers, data)

    def _export_active_strategies_sheet(self, wb):
        """导出活跃策略"""
        ws = wb.create_sheet("活跃策略")

        cursor = self.conn.cursor()
        cursor.execute(f"""
            SELECT {self.FIELD_STRATEGY_ID}, {self.FIELD_INST_ID}, {self.FIELD_STRATEGY_TYPE},
                   {self.FIELD_START_TIME}, {self.FIELD_TOTAL_MARGIN_USD}, {self.FIELD_TOTAL_CONTRACTS},
                   {self.FIELD_AVG_ENTRY_PRICE}, {self.FIELD_TP_PRICE}, {self.FIELD_SL_PRICE}, {self.FIELD_LIQ_PRICE}
            FROM strategies
            WHERE {self.FIELD_STATUS} = 'RUNNING'
            ORDER BY {self.FIELD_START_TIME} DESC
        """)

        rows = cursor.fetchall()

        headers = ["策略ID", "交易对", "策略类型", "开始时间", "总保证金", "总张数", "均价", "止盈价", "止损价", "爆仓价"]
        data = []

        for row in rows:
            data.append([
                row[self.FIELD_STRATEGY_ID],
                row[self.FIELD_INST_ID],
                row[self.FIELD_STRATEGY_TYPE],
                row[self.FIELD_START_TIME],
                row[self.FIELD_TOTAL_MARGIN_USD],
                row[self.FIELD_TOTAL_CONTRACTS],
                row[self.FIELD_AVG_ENTRY_PRICE],
                row[self.FIELD_TP_PRICE],
                row[self.FIELD_SL_PRICE],
                row[self.FIELD_LIQ_PRICE]
            ])

        self._format_sheet(ws, headers, data)

    def _export_tp_sl_orders_sheet(self, wb):
        """导出止盈止损单"""
        ws = wb.create_sheet("止盈止损单")

        cursor = self.conn.cursor()
        cursor.execute(f"""
            SELECT {self.FIELD_ID}, {self.FIELD_STRATEGY_ID}, {self.FIELD_INST_ID}, {self.FIELD_TP_TRIGGER_PX},
                   {self.FIELD_SL_TRIGGER_PX}, {self.FIELD_SZ}, {self.FIELD_STATUS}, {self.FIELD_CREATED_TIME},
                   {self.FIELD_TRIGGERED_TIME}, {self.FIELD_TRIGGER_TYPE}, {self.FIELD_EXIT_PRICE}
            FROM tp_sl_orders
            ORDER BY {self.FIELD_CREATED_TIME} DESC
            LIMIT 100
        """)

        rows = cursor.fetchall()

        headers = ["ID", "策略ID", "交易对", "止盈价", "止损价", "数量", "状态", "创建时间", "触发时间", "触发类型", "退出价格"]
        data = []

        for row in rows:
            data.append([
                row[self.FIELD_ID],
                row[self.FIELD_STRATEGY_ID],
                row[self.FIELD_INST_ID],
                row[self.FIELD_TP_TRIGGER_PX],
                row[self.FIELD_SL_TRIGGER_PX],
                row[self.FIELD_SZ],
                row[self.FIELD_STATUS],
                row[self.FIELD_CREATED_TIME],
                self._format_value(row[self.FIELD_TRIGGERED_TIME]),
                self._format_value(row[self.FIELD_TRIGGER_TYPE]),
                self._format_value(row[self.FIELD_EXIT_PRICE])
            ])

        self._format_sheet(ws, headers, data)

    def _export_orders_sheet(self, wb):
        """导出订单记录"""
        ws = wb.create_sheet("订单记录")

        cursor = self.conn.cursor()
        cursor.execute(f"""
            SELECT {self.FIELD_ID}, {self.FIELD_ORDER_ID}, {self.FIELD_STRATEGY_ID}, {self.FIELD_INST_ID},
                   {self.FIELD_SIDE}, {self.FIELD_POS_SIDE}, {self.FIELD_ORD_TYPE}, {self.FIELD_PX},
                   {self.FIELD_SZ}, {self.FIELD_FILLED_SZ}, {self.FIELD_AVG_PX}, {self.FIELD_STATE}, {self.FIELD_SOURCE},
                   {self.FIELD_CREATED_TIME}
            FROM orders
            ORDER BY {self.FIELD_CREATED_TIME} DESC
            LIMIT 200
        """)

        rows = cursor.fetchall()

        headers = ["ID", "订单ID", "策略ID", "交易对", "方向", "持仓方向", "订单类型", "价格", "数量", "成交数量", "均价", "状态", "来源", "创建时间"]
        data = []

        for row in rows:
            data.append([
                row[self.FIELD_ID],
                row[self.FIELD_ORDER_ID],
                row[self.FIELD_STRATEGY_ID],
                row[self.FIELD_INST_ID],
                row[self.FIELD_SIDE],
                row[self.FIELD_POS_SIDE],
                row[self.FIELD_ORD_TYPE],
                row[self.FIELD_PX],
                row[self.FIELD_SZ],
                row[self.FIELD_FILLED_SZ],
                row[self.FIELD_AVG_PX],
                row[self.FIELD_STATE],
                row[self.FIELD_SOURCE],
                row[self.FIELD_CREATED_TIME]
            ])

        self._format_sheet(ws, headers, data)

    def _export_action_logs_sheet(self, wb):
        """导出行为日志"""
        ws = wb.create_sheet("行为日志")

        cursor = self.conn.cursor()
        cursor.execute(f"""
            SELECT {self.FIELD_ID}, {self.FIELD_TIMESTAMP}, {self.FIELD_STRATEGY_ID}, {self.FIELD_INST_ID},
                   {self.FIELD_ACTION_TYPE}, {self.FIELD_ACTION_DETAIL}
            FROM action_logs
            ORDER BY {self.FIELD_TIMESTAMP} DESC
            LIMIT 100
        """)

        rows = cursor.fetchall()

        headers = ["ID", "时间", "策略ID", "交易对", "行为类型", "详情"]
        data = []

        for row in rows:
            detail = row[self.FIELD_ACTION_DETAIL] or ""
            # 截断过长的详情（确保 detail 不为 None）
            if len(detail) > 200:
                detail = detail[:200] + "..."

            data.append([
                row[self.FIELD_ID],
                row[self.FIELD_TIMESTAMP],
                self._format_value(row[self.FIELD_STRATEGY_ID]),
                self._format_value(row[self.FIELD_INST_ID]),
                row[self.FIELD_ACTION_TYPE],
                detail
            ])

        self._format_sheet(ws, headers, data)

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()
            self.conn = None


@contextmanager
def create_excel_exporter(db_path: str = "data/trading_history.db"):
    """
    创建 Excel 导出器的上下文管理器

    :param db_path: 数据库文件路径
    """
    exporter = ExcelExporter(db_path)
    try:
        yield exporter
    finally:
        exporter.close()


def export_excel(db_path: str = "data/trading_history.db", output_path: str = None) -> str:
    """
    导出数据到 Excel 文件（便捷函数）

    :param db_path: 数据库文件路径
    :param output_path: 输出文件路径
    :return: Excel 文件路径
    """
    try:
        with create_excel_exporter(db_path) as exporter:
            return exporter.export_to_excel(output_path)
    except Exception as e:
        raise e
